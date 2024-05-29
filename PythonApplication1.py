
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select
import time
from datetime import datetime, date
import openpyxl
from openpyxl.cell import Cell
import re

def contains_monetary_value(input_string):
    # Define regular expressions for different monetary formats
    currency_pattern = r'\$[\d.,]+|\b\d+\s*(?:dollars|USD)\b'

    # Search for monetary values in the input string
    match = re.search(currency_pattern, input_string, flags=re.IGNORECASE)

    return bool(match)

def append_year_to_string(input_string):
    if len(input_string) < 11:
        current_year = str(datetime.now().year)
        return input_string + ", " + current_year
    else:
        return input_string

def compare_dates(input_date):    
        if input_date.endswith('ago'):
            return 0
        else:
            #Append current yaer if input date contains none
            input_date = append_year_to_string(input_date)
            # Try 1st date format    
            try:
                # Parse the input date (assuming it's in the format 'May 12, 2023')
                input_date_obj = datetime.strptime(input_date, '%B %d, %Y')
        
                # Get the current date
                current_date_obj = datetime.now()
        
                # Calculate the difference in months
                months_diff = (current_date_obj.year - input_date_obj.year) * 12 + current_date_obj.month - input_date_obj.month
        
                return months_diff
            # Try second date format
            except ValueError:
                try:
                    # Parse the input date (assuming it's in the format 'Dec. 12, 2023')
                    input_date_obj = datetime.strptime(input_date, '%b. %d, %Y')
        
                    # Get the current date
                    current_date_obj = datetime.now()
        
                    # Calculate the difference in months
                    months_diff = (current_date_obj.year - input_date_obj.year) * 12 + current_date_obj.month - input_date_obj.month
        
                    return months_diff
                
                except ValueError:
                    return 'Invalid Date Format'                          

chosen_section = "Arts"
user_input = "Trump"
months_selected = 1

# Initialize the Chrome WebDriver
driver = webdriver.Chrome()

# Open New York Times website
driver.get("https://www.nytimes.com/")

# Handle cookies
try:
    reject_button = driver.find_element(By.CLASS_NAME, "fides-banner-button fides-banner-button-primary fides-reject-all-button")
    reject_button.click()
except NoSuchElementException:
    pass

time.sleep(1)

# Click search icon for search bar dropdown
search_icon = driver.find_element(By.CLASS_NAME, "css-etfx3o")
search_icon.click()

# Enter text into search bar
search_box = driver.find_element(By.CLASS_NAME, "css-1u4s13l")
search_box.send_keys(user_input)

# Submit search
submit_button = driver.find_element(By.CLASS_NAME, "css-1gudca6.e1iflr852")
submit_button.click()

# Repating wait loop to check if website loaded
time.sleep(2)

# Select news category
section_selector_button = driver.find_element(By.CLASS_NAME, "css-4d08fs")
section_selector_button.click()

for element in driver.find_elements(By.CLASS_NAME, "css-1qtb2wd"):
    if element.text.startswith(chosen_section) == True:
        element.click()
        
section_selector_button.click()

time.sleep(3)    

# Select latest news, use select method
ddelement = driver.find_element(By.CLASS_NAME, "css-v7it2b")
select = Select(ddelement)
select.select_by_visible_text('Sort by Newest')

time.sleep(3) 
    
# fetch article elements currently visible
article_dates = driver.find_elements(By.CSS_SELECTOR, "span[class='css-17ubb9w']")

# while latest date is more recent than specified months
while compare_dates(article_dates[-1].text) < months_selected:
    # click 'show more'
    driver.find_element(By.CSS_SELECTOR, "button[data-testid='search-show-more-button']").click()
    # print(compare_dates(articles[-1].text))
    # get updated article elements
    article_dates = driver.find_elements(By.CSS_SELECTOR, "span[class='css-17ubb9w']")
    time.sleep(2) 
    
    
# Create Excel file
wb = openpyxl.Workbook()
sheet = wb.active

# Create headers for table in Excel
sheet.cell(row = 1, column = 1).value = 'Title'
sheet.cell(row = 1, column = 2).value = 'Date'
sheet.cell(row = 1, column = 3).value = 'Description'
sheet.cell(row = 1, column = 4).value = 'Picture File Name'
sheet.cell(row = 1, column = 5).value = 'Phrase Count'
sheet.cell(row = 1, column = 6).value = 'Article Contains Monetary Value'

row_num = 2

# Extract requested information about aricles
articles = driver.find_elements(By.CSS_SELECTOR, "li[class='css-1l4w6pd']")
for article in articles:
    
    if compare_dates(article.find_element(By.CSS_SELECTOR, "span[class='css-17ubb9w']").text) < months_selected:
        
        # Title
        title = article.find_element(By.CSS_SELECTOR, "h4[class='css-nsjm9t']").text
        sheet.cell(row = row_num, column = 1).value = title
        # Date
        sheet.cell(row = row_num, column = 2).value = article.find_element(By.CSS_SELECTOR, "span[class='css-17ubb9w']").text
        # Description
        description = article.find_element(By.CSS_SELECTOR, "p[class='css-16nhkrn']").text
        sheet.cell(row = row_num, column = 3).value = description
        # Save picture
        img = article.find_element(By.CSS_SELECTOR, "img[class='css-rq4mmj']")
        img.screenshot("D:\\Extra Studies\\Programming\\Thoughtful\\article_" + str(row_num-1) + ".png")
        # Picture filename
        sheet.cell(row = row_num, column = 4).value = "article_" + str(row_num-1) + ".png"
        # Search count
        my_string = title + description
        count = my_string.count(user_input)
        sheet.cell(row = row_num, column = 5).value = count
        # If the title contains monetary value
        sheet.cell(row = row_num, column = 6).value = contains_monetary_value(my_string)
        
        
        row_num = row_num + 1
        
# Format Excel File
tab = openpyxl.worksheet.table.Table(displayName="Table1", ref="A1:F" + str(row_num))

# Add a default style with striped rows and banded columns
style = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style
sheet.add_table(tab)

# Adjust columns A to F width
sheet.column_dimensions['A'].width = 30
sheet.column_dimensions['B'].width = 10
sheet.column_dimensions['C'].width = 40
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 20
sheet.column_dimensions['F'].width = 20
    
# Adjust row heights
for row in range(2, row_num):
        sheet.row_dimensions[row].height = 70
        
# Set wrap text
for row in sheet.iter_rows():  
    for cell in row:      
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True,vertical='top')
    
# Save file
# Incorporate Dynamic Naming Scheme
file_name =  user_input + '_' + chosen_section +  '_' + str(months_selected) +  '_Months'    
wb.save("D:\\Extra Studies\\Programming\\Thoughtful\\" + file_name + ".xlsx")

time.sleep(2)

driver.quit()